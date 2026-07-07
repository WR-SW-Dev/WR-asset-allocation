"""Morningstar Direct index-return ingestion for the Asset Allocation Study Model.

Export-based (file) ingestion of a broad index basket into a clean, auditable
long-format return store. Complements — does not replace — the live-daily-feed CMA
adapter in :mod:`aa_model.assumptions.benchmark_calibration`.

What it reads
-------------
The primary source is the Morningstar Direct "Common Indices" export
("Index Returns - <date>.xlsx"): a *cross-sectional trailing-return snapshot*,
one row per index, columns per trailing window (1 Mo / 3 Mo / 6 Mo / 1 Yr /
Annlzd 3-15 Yr / Inception). It also parses two generic time-series layouts —
wide (dates in column 0, one column per index) and long (date/index/value) — so
the same importer can absorb hand-built monthly series.

Canonical store (long by horizon)
---------------------------------
One row per (``index_key``, ``date``, ``horizon``). Columns are fixed by
:data:`morningstar_schemas.NORMALIZED_COLUMNS`. The ``horizon == "1M"`` slice at a
month-end date is the canonical *monthly total return* series that CMA / Monte
Carlo / study exhibits consume; longer horizons are retained for realized-stat
exhibits. Returns are stored as **decimals** (``1.25`` pct -> ``0.0125``).

Contract
--------
* Pure / deterministic: same input bytes + configs -> same ``IngestionResult``
  (modulo ``fetched_at_utc`` provenance, which is intentionally wall-clock).
* Read-only: workbooks opened ``read_only=True, data_only=True, keep_links=False``.
* Loud: ambiguous layouts and mapping violations raise, they do not guess.
* No live network access; no credentials; no writes unless the caller asks.
"""

from __future__ import annotations

import csv
import datetime as _dt
import hashlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

from aa_model.ingestion.morningstar_schemas import (
    BASE_CURRENCY_HEADER,
    CANONICAL_MONTHLY_HORIZON,
    HORIZONS,
    NAME_HEADER,
    NORMALIZED_COLUMNS,
    RETURN_DATE_HEADER,
    SOURCE_NAME,
    SUMMARY_BLOCK_SENTINEL,
    SUMMARY_ROW_LABELS,
    TOTAL_RETURN_HEADER_TO_HORIZON,
    AssetClassIndexMapConfig,
    IndexUniverseConfig,
)

# ---- tuning constants ------------------------------------------------------

# |return_decimal| above this is flagged EXTREME_RETURN (audit hint, not a drop).
EXTREME_RETURN_THRESHOLD = 0.60
# Horizons whose absence marks an index as SHORT_HISTORY.
_LONG_HORIZONS = ("5Y_ann", "10Y_ann")
# Currency label normalization (Morningstar spells USD as "US Dollar").
_CURRENCY_ALIASES = {"us dollar": "USD", "usd": "USD", "$": "USD"}

_DATE_HEADER_TOKENS = {
    "",
    "date",
    "return date (mo-end)",
    "return_date",
    "month",
    "month_end",
    "period",
    "as_of",
    "asof",
}
_INDEX_HEADER_TOKENS = {"index", "index_key", "name", "benchmark", "series"}
_VALUE_HEADER_TOKENS = {
    "value",
    "return",
    "return_pct",
    "return_decimal",
    "total_return",
    "ret",
    "total ret",
}

_MAX_HEADER_SCAN_ROWS = 40


class MorningstarIngestError(ValueError):
    """Raised on ambiguous layouts or mapping-contract violations."""


# ---- config loaders --------------------------------------------------------


def load_index_universe(path: Path | str) -> IndexUniverseConfig:
    raw = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    return IndexUniverseConfig.model_validate(raw)


def load_asset_class_map(path: Path | str) -> AssetClassIndexMapConfig:
    raw = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    return AssetClassIndexMapConfig.model_validate(raw)


def validate_asset_class_map(
    map_cfg: AssetClassIndexMapConfig, universe: IndexUniverseConfig
) -> None:
    """Every index_key referenced by the asset-class map must exist in the universe."""
    known = set(universe.by_index_key())
    referenced = map_cfg.referenced_index_keys()
    unknown = sorted(referenced - known)
    if unknown:
        raise MorningstarIngestError(
            f"asset_class_index_map references unknown index_key(s) not in the "
            f"index universe: {unknown}"
        )


# ---- low-level readers -----------------------------------------------------


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    h.update(Path(path).read_bytes())
    return f"sha256:{h.hexdigest()}"


def _read_xlsx_rows(path: Path, sheet: str | None) -> tuple[list[list[Any]], str]:
    """Return (rows, sheet_name). Read-only, cached-value, no external links."""
    import openpyxl  # lazy: keep the module import-light

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise MorningstarIngestError(
                    f"sheet {sheet!r} not found; available: {wb.sheetnames}"
                )
            name = sheet
        elif "Common Indices" in wb.sheetnames:
            name = "Common Indices"
        else:
            name = wb.sheetnames[0]
        ws = wb[name]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        return rows, name
    finally:
        wb.close()


def _read_csv_rows(path: Path) -> list[list[Any]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as f:
        return [list(r) for r in csv.reader(f)]


def _read_rows(path: Path, sheet: str | None) -> tuple[list[list[Any]], str | None]:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path, sheet)
    if suffix == ".csv":
        return _read_csv_rows(path), None
    raise MorningstarIngestError(
        f"unsupported input extension {suffix!r}; expected .xlsx/.xlsm/.csv"
    )


# ---- helpers ---------------------------------------------------------------


def _norm_label(x: Any) -> str:
    return "" if x is None else str(x).strip()


def _to_float(x: Any) -> float | None:
    """Blank -> None; otherwise a float (raises loudly on garbage)."""
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip().replace(",", "")
        if s == "" or s in {"-", "--", "N/A", "NA", "nan", "NaN"}:
            return None
        return float(s)
    if isinstance(x, int | float):
        f = float(x)
        return None if pd.isna(f) else f
    raise MorningstarIngestError(f"non-numeric return cell: {x!r}")


def _parse_date(x: Any) -> pd.Timestamp | None:
    if x is None or (isinstance(x, str) and x.strip() == ""):
        return None
    ts = pd.to_datetime(x, errors="coerce")
    if pd.isna(ts):
        return None
    return pd.Timestamp(ts).normalize()


def _is_month_end(ts: pd.Timestamp) -> bool:
    return bool(ts.is_month_end)


def _snap_month_end(ts: pd.Timestamp) -> pd.Timestamp:
    return (ts + pd.offsets.MonthEnd(0)).normalize()


def _norm_currency(x: Any) -> str | None:
    s = _norm_label(x).lower()
    if s == "":
        return None
    return _CURRENCY_ALIASES.get(s, _norm_label(x))


# ---- layout detection ------------------------------------------------------


def _locate_header(rows: list[list[Any]]) -> tuple[int, str]:
    """Find (header_row_index, layout). Raise loudly if ambiguous.

    layout in {"morningstar_common_indices", "long", "wide"}.
    Scans up to _MAX_HEADER_SCAN_ROWS to tolerate leading title/blank rows
    (Morningstar and hand-built exports both carry header rows above the table).
    """
    for i, row in enumerate(rows[:_MAX_HEADER_SCAN_ROWS]):
        labels = [_norm_label(c) for c in row]
        low = [c.lower() for c in labels]
        low_set = set(low)
        nonempty = [c for c in labels if c != ""]
        if len(nonempty) < 2:
            continue

        # (1) Morningstar "Common Indices" trailing-return matrix — most specific.
        if NAME_HEADER in labels and RETURN_DATE_HEADER in labels:
            if any(h in labels for h in TOTAL_RETURN_HEADER_TO_HORIZON):
                return i, "morningstar_common_indices"

        # (2) Long: a date-ish, an index-ish, and a value-ish column.
        has_date = bool(low_set & (_DATE_HEADER_TOKENS - {""}))
        has_index = bool(low_set & _INDEX_HEADER_TOKENS)
        has_value = bool(low_set & _VALUE_HEADER_TOKENS)
        if has_date and has_index and has_value:
            return i, "long"

        # (3) Wide time series: column 0 is a date column; the rest are index names.
        col0 = low[0] if low else ""
        if col0 in _DATE_HEADER_TOKENS and len(nonempty) >= 2:
            # confirm the next non-empty row's first cell parses as a date
            for nxt in rows[i + 1 : i + 4]:
                if nxt and _parse_date(nxt[0]) is not None:
                    return i, "wide"

    sample = [
        [_norm_label(c) for c in r][:8] for r in rows[: min(6, len(rows))]
    ]
    raise MorningstarIngestError(
        "could not detect input layout (morningstar_common_indices / long / wide). "
        f"First rows seen: {sample}. For wide format name column 0 'Date'; for long "
        "format include date/index/value columns; for a Morningstar export keep the "
        f"native '{NAME_HEADER}' + '{RETURN_DATE_HEADER}' header."
    )


# ---- per-layout parsers ----------------------------------------------------


@dataclass(frozen=True)
class _RawObs:
    """One raw (pre-normalization) observation."""

    display_name: str
    date: pd.Timestamp | None
    raw_date: Any
    horizon: str
    value_pct: float | None
    source_field: str
    currency: Any = None


def _parse_morningstar(rows: list[list[Any]], header_i: int) -> list[_RawObs]:
    header = [_norm_label(c) for c in rows[header_i]]
    col = {h: idx for idx, h in enumerate(header)}
    name_c = col[NAME_HEADER]
    date_c = col[RETURN_DATE_HEADER]
    cur_c = col.get(BASE_CURRENCY_HEADER)
    horizon_cols = {
        h: col[label]
        for label, h in TOTAL_RETURN_HEADER_TO_HORIZON.items()
        if label in col
    }
    inv_field = {v: k for k, v in TOTAL_RETURN_HEADER_TO_HORIZON.items()}

    out: list[_RawObs] = []
    for row in rows[header_i + 1 :]:
        name = _norm_label(row[name_c]) if name_c < len(row) else ""
        low = name.lower()
        if name == "":
            continue  # blank separator row
        if low == SUMMARY_BLOCK_SENTINEL or low in SUMMARY_ROW_LABELS:
            break  # summary block starts here — stop ingesting
        raw_date = row[date_c] if date_c < len(row) else None
        date = _parse_date(raw_date)
        cur = row[cur_c] if (cur_c is not None and cur_c < len(row)) else None
        for horizon, cidx in horizon_cols.items():
            raw_val = row[cidx] if cidx < len(row) else None
            out.append(
                _RawObs(
                    display_name=name,
                    date=date,
                    raw_date=raw_date,
                    horizon=horizon,
                    value_pct=_to_float(raw_val),
                    source_field=inv_field[horizon],
                    currency=cur,
                )
            )
    return out


def _parse_wide(rows: list[list[Any]], header_i: int) -> list[_RawObs]:
    header = [_norm_label(c) for c in rows[header_i]]
    name_cols = [(idx, h) for idx, h in enumerate(header) if idx != 0 and h != ""]
    out: list[_RawObs] = []
    for row in rows[header_i + 1 :]:
        if not row or all(_norm_label(c) == "" for c in row):
            continue  # blank/footer row
        raw_date = row[0] if row else None
        date = _parse_date(raw_date)
        if date is None:
            continue  # footer text under the table
        for idx, name in name_cols:
            raw_val = row[idx] if idx < len(row) else None
            out.append(
                _RawObs(
                    display_name=name,
                    date=date,
                    raw_date=raw_date,
                    horizon=CANONICAL_MONTHLY_HORIZON,
                    value_pct=_to_float(raw_val),
                    source_field=f"wide:{name}",
                )
            )
    return out


def _parse_long(rows: list[list[Any]], header_i: int) -> list[_RawObs]:
    header = [_norm_label(c) for c in rows[header_i]]
    low = [h.lower() for h in header]

    def _find(tokens: set[str]) -> int:
        for idx, h in enumerate(low):
            if h in tokens:
                return idx
        raise MorningstarIngestError(f"long layout missing a column in {sorted(tokens)}")

    date_c = _find(_DATE_HEADER_TOKENS - {""})
    index_c = _find(_INDEX_HEADER_TOKENS)
    value_c = _find(_VALUE_HEADER_TOKENS)
    value_field = header[value_c]
    out: list[_RawObs] = []
    for row in rows[header_i + 1 :]:
        if not row or all(_norm_label(c) == "" for c in row):
            continue
        name = _norm_label(row[index_c]) if index_c < len(row) else ""
        if name == "":
            continue
        raw_date = row[date_c] if date_c < len(row) else None
        date = _parse_date(raw_date)
        raw_val = row[value_c] if value_c < len(row) else None
        out.append(
            _RawObs(
                display_name=name,
                date=date,
                raw_date=raw_date,
                horizon=CANONICAL_MONTHLY_HORIZON,
                value_pct=_to_float(raw_val),
                source_field=f"long:{value_field}",
            )
        )
    return out


_PARSERS = {
    "morningstar_common_indices": _parse_morningstar,
    "wide": _parse_wide,
    "long": _parse_long,
}


# ---- normalization ---------------------------------------------------------


@dataclass
class IngestionResult:
    normalized: pd.DataFrame
    coverage: pd.DataFrame
    meta: dict[str, Any]
    unmapped_names: list[str] = field(default_factory=list)
    missing_configured: list[str] = field(default_factory=list)


def _scale_factor(value_scale: str) -> float:
    if value_scale == "percent":
        return 0.01  # applied EXACTLY once here — no double-division elsewhere
    if value_scale == "decimal":
        return 1.0
    raise MorningstarIngestError(f"value_scale must be 'percent' or 'decimal'; got {value_scale!r}")


def _normalize(
    raws: list[_RawObs],
    universe: IndexUniverseConfig,
    *,
    value_scale: str,
    asof_date: pd.Timestamp | None,
    fetched_at_utc: str,
) -> tuple[pd.DataFrame, pd.Timestamp]:
    by_name = universe.by_display_name()
    scale = _scale_factor(value_scale)

    # Determine the ingestion as-of: explicit override, else the modal (most
    # common) non-null observation date among MAPPED rows.
    mapped_dates = [
        o.date for o in raws if o.display_name in by_name and o.date is not None
    ]
    if asof_date is not None:
        asof = asof_date.normalize()
    elif mapped_dates:
        asof = pd.Series(mapped_dates).mode().iloc[0].normalize()
    else:
        asof = pd.Timestamp(fetched_at_utc[:10]).normalize()

    # Index-level SHORT_HISTORY: any mapped index missing a long-horizon value.
    long_present: dict[str, set[str]] = {}
    for o in raws:
        if o.display_name in by_name and o.value_pct is not None:
            long_present.setdefault(o.display_name, set()).add(o.horizon)
    short_history_names = {
        name
        for name in long_present
        if not any(h in long_present[name] for h in _LONG_HORIZONS)
    }

    records: list[dict[str, Any]] = []
    seen: set[tuple[str, Any, str]] = set()
    dup_keys: set[tuple[str, Any, str]] = set()
    for o in raws:
        entry = by_name.get(o.display_name)
        if entry is None:
            continue  # unmapped names are handled by validate_mapping
        # date normalization policy: snap non-month-end to month-end + flag.
        flags: set[str] = set()
        date = o.date
        if date is not None and not _is_month_end(date):
            flags.add("NON_MONTH_END_DATE")
            date = _snap_month_end(date)

        date_iso = None if date is None else date.date().isoformat()
        key = (entry.index_key, date_iso, o.horizon)
        if key in seen:
            dup_keys.add(key)
            flags.add("DUPLICATE_DATE")
        seen.add(key)

        value = None if o.value_pct is None else round(o.value_pct * scale, 10)

        if value is None:
            flags.add("MISSING_MONTH" if o.horizon == CANONICAL_MONTHLY_HORIZON else "SHORT_HISTORY")
        else:
            if abs(value) > EXTREME_RETURN_THRESHOLD:
                flags.add("EXTREME_RETURN")
        if date is not None and date < asof:
            flags.add("STALE_SERIES")
        if o.display_name in short_history_names:
            flags.add("SHORT_HISTORY")
        row_cur = _norm_currency(o.currency)
        if row_cur is not None and row_cur != entry.currency:
            flags.add("CURRENCY_MISMATCH")
        if entry.return_type == "unknown":
            flags.add("RETURN_TYPE_UNKNOWN")

        quality = "OK" if not flags else "|".join(sorted(flags))
        records.append(
            {
                "date": None if date is None else date.date(),
                "index_key": entry.index_key,
                "horizon": o.horizon,
                "return_decimal": value,
                "annualized": HORIZONS.get(o.horizon, False),
                "level": None,
                "currency": entry.currency,
                "source": SOURCE_NAME,
                "source_field": o.source_field,
                "return_type": entry.return_type,
                "frequency": entry.frequency,
                "asof_date": asof.date(),
                "fetched_at_utc": fetched_at_utc,
                "vendor_id": entry.morningstar_id,
                "quality_flag": quality,
                "notes": entry.notes,
            }
        )

    df = pd.DataFrame.from_records(records, columns=list(NORMALIZED_COLUMNS))
    # Second pass: mark BOTH members of any duplicate (index_key,date,horizon).
    if dup_keys:
        def _mark(r: pd.Series) -> str:
            d_iso = None if r["date"] is None else pd.Timestamp(r["date"]).date().isoformat()
            k = (r["index_key"], d_iso, r["horizon"])
            if k in dup_keys and "DUPLICATE_DATE" not in str(r["quality_flag"]):
                base = set() if r["quality_flag"] == "OK" else set(str(r["quality_flag"]).split("|"))
                base.add("DUPLICATE_DATE")
                return "|".join(sorted(base))
            return r["quality_flag"]

        df["quality_flag"] = df.apply(_mark, axis=1)
    return df, asof


# ---- mapping validation ----------------------------------------------------


def validate_mapping(
    workbook_names: list[str],
    universe: IndexUniverseConfig,
    *,
    allow_unmapped: bool,
    allow_missing_configured: bool,
) -> tuple[list[str], list[str]]:
    """Return (unmapped_names, missing_configured); raise unless the matching
    ``allow_*`` flag is set. ``unmapped_names`` are workbook rows with no
    universe entry; ``missing_configured`` are universe entries absent from the
    workbook."""
    by_name = universe.by_display_name()
    wb_set = list(dict.fromkeys(workbook_names))  # preserve order, dedupe
    unmapped = sorted({n for n in wb_set if n not in by_name})
    missing = sorted({n for n in by_name if n not in set(wb_set)})

    if unmapped and not allow_unmapped:
        raise MorningstarIngestError(
            f"{len(unmapped)} workbook index name(s) are not in "
            f"morningstar_index_universe.yaml: {unmapped}. Add them to the universe "
            "or pass --allow-unmapped."
        )
    if missing and not allow_missing_configured:
        raise MorningstarIngestError(
            f"{len(missing)} configured index/indices are missing from the workbook: "
            f"{missing}. Fix the export or pass --allow-missing-configured."
        )
    return unmapped, missing


# ---- coverage report -------------------------------------------------------


def build_coverage_report(
    normalized: pd.DataFrame,
    universe: IndexUniverseConfig,
    *,
    asof_date: pd.Timestamp,
    present_names: set[str],
) -> pd.DataFrame:
    """One row per CONFIGURED index (every index in the universe appears)."""
    rows: list[dict[str, Any]] = []
    for entry in universe.indices:
        key = entry.index_key
        sub = (
            normalized[normalized["index_key"] == key]
            if not normalized.empty
            else normalized
        )
        present = entry.display_name in present_names
        obs_dates = sorted(
            {r for r in sub["date"].tolist() if r is not None}
        ) if not sub.empty else []
        monthly = sub[sub["horizon"] == CANONICAL_MONTHLY_HORIZON] if not sub.empty else sub
        monthly_valid = (
            monthly[monthly["return_decimal"].notna()] if not monthly.empty else monthly
        )
        return_date = obs_dates[-1] if obs_dates else None
        n_monthly = int(len(monthly_valid))
        all_flags: set[str] = set()
        for q in (sub["quality_flag"].tolist() if not sub.empty else []):
            if q and q != "OK":
                all_flags.update(str(q).split("|"))
        stale = "STALE_SERIES" in all_flags or (
            return_date is not None and pd.Timestamp(return_date) < asof_date
        )
        # missing months: for a single snapshot this is 0/1 of the modal month.
        expected_monthly = 1 if entry.frequency == "monthly" else 0
        missing_months = max(expected_monthly - n_monthly, 0)
        rows.append(
            {
                "index_key": key,
                "display_name": entry.display_name,
                "asset_class": entry.asset_class,
                "sub_asset_class": entry.sub_asset_class,
                "model_role": entry.model_role,
                "present_in_workbook": present,
                "return_date": return_date,
                "first_observation_date": obs_dates[0] if obs_dates else None,
                "last_observation_date": return_date,
                "n_monthly_observations": n_monthly,
                "n_horizons": int(sub["horizon"].nunique()) if not sub.empty else 0,
                "missing_months": missing_months,
                "stale_series": bool(stale),
                "return_type": entry.return_type,
                "currency": entry.currency,
                "currency_flag": "CURRENCY_MISMATCH" in all_flags,
                "return_type_flag": "RETURN_TYPE_UNKNOWN" in all_flags,
                "short_history_flag": "SHORT_HISTORY" in all_flags,
                "quality_flags": "|".join(sorted(all_flags)) if all_flags else "OK",
                "notes": entry.notes,
            }
        )
    return pd.DataFrame.from_records(rows)


# ---- orchestration ---------------------------------------------------------


def run_ingestion(
    input_path: Path | str,
    *,
    universe_path: Path | str,
    asset_class_map_path: Path | str | None = None,
    sheet: str | None = "Common Indices",
    asof: str | None = None,
    value_scale: str = "percent",
    allow_unmapped: bool = False,
    allow_missing_configured: bool = False,
) -> IngestionResult:
    """Parse -> validate mapping -> normalize -> coverage. Pure; performs no writes."""
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"input not found: {input_path}")

    universe = load_index_universe(universe_path)
    if asset_class_map_path is not None:
        map_cfg = load_asset_class_map(asset_class_map_path)
        validate_asset_class_map(map_cfg, universe)

    rows, sheet_name = _read_rows(input_path, sheet if input_path.suffix.lower() != ".csv" else None)
    header_i, layout = _locate_header(rows)
    raws = _PARSERS[layout](rows, header_i)

    workbook_names = [o.display_name for o in raws]
    unmapped, missing = validate_mapping(
        workbook_names,
        universe,
        allow_unmapped=allow_unmapped,
        allow_missing_configured=allow_missing_configured,
    )

    fetched_at_utc = _dt.datetime.now(_dt.UTC).isoformat(timespec="seconds")
    asof_ts = _parse_date(asof) if asof else None
    normalized, asof_used = _normalize(
        raws,
        universe,
        value_scale=value_scale,
        asof_date=asof_ts,
        fetched_at_utc=fetched_at_utc,
    )
    present_names = {n for n in workbook_names if n in universe.by_display_name()}
    coverage = build_coverage_report(
        normalized, universe, asof_date=asof_used, present_names=present_names
    )

    meta = {
        "input_path": str(input_path),
        "input_sha256": _sha256(input_path),
        "sheet": sheet_name,
        "layout": layout,
        "value_scale": value_scale,
        "asof_date": asof_used.date().isoformat(),
        "fetched_at_utc": fetched_at_utc,
        "n_source_rows": len({n for n in workbook_names}),
        "n_normalized_rows": int(len(normalized)),
        "n_unmapped": len(unmapped),
        "n_missing_configured": len(missing),
        "source": SOURCE_NAME,
    }
    return IngestionResult(
        normalized=normalized,
        coverage=coverage,
        meta=meta,
        unmapped_names=unmapped,
        missing_configured=missing,
    )


# ---- model consumption (Step 6) --------------------------------------------


def load_normalized_returns(path: Path | str) -> pd.DataFrame:
    """Load a previously written normalized store (.parquet or .csv)."""
    p = Path(path)
    if p.suffix.lower() == ".parquet":
        return pd.read_parquet(p)
    return pd.read_csv(p)


def monthly_return_series(df: pd.DataFrame, index_key: str) -> pd.Series:
    """Canonical monthly total-return series for one index (horizon == 1M),
    indexed by date, decimals, NaNs dropped. Empty Series if absent."""
    sub = df[(df["index_key"] == index_key) & (df["horizon"] == CANONICAL_MONTHLY_HORIZON)]
    sub = sub[sub["return_decimal"].notna()]
    if sub.empty:
        return pd.Series(dtype=float, name=index_key)
    s = pd.Series(
        sub["return_decimal"].to_numpy(dtype=float),
        index=pd.to_datetime(sub["date"]),
        name=index_key,
    ).sort_index()
    return s


def trailing_return(df: pd.DataFrame, index_key: str, horizon: str) -> float | None:
    """The stored trailing return (decimal) for (index_key, horizon), latest date."""
    if horizon not in HORIZONS:
        raise MorningstarIngestError(f"unknown horizon {horizon!r}; valid: {sorted(HORIZONS)}")
    sub = df[(df["index_key"] == index_key) & (df["horizon"] == horizon)]
    sub = sub[sub["return_decimal"].notna()]
    if sub.empty:
        return None
    latest = sub.sort_values("date").iloc[-1]
    return float(latest["return_decimal"])


def index_keys_for_asset_class(
    map_cfg: AssetClassIndexMapConfig,
    asset_class: str,
    sub_asset_class: str | None = None,
) -> dict[str, Any]:
    """Resolve the primary/secondary/fallback index_keys for an asset class.

    Returns the matching map entry as a dict, or raises if not found. Callers
    (CMA calibration, Monte Carlo assumption sourcing, exhibits) use this to pick
    which normalized series represents the class.
    """
    for e in map_cfg.asset_classes:
        if e.asset_class == asset_class and (
            sub_asset_class is None or e.sub_asset_class == sub_asset_class
        ):
            return {
                "primary_index_key": e.primary_index_key,
                "secondary_index_keys": list(e.secondary_index_keys),
                "fallback_index_key": e.fallback_index_key,
                "model_usage": e.model_usage,
                "min_history_months": e.min_history_months,
                "requires_approval": e.requires_approval,
            }
    raise MorningstarIngestError(
        f"no asset-class map entry for asset_class={asset_class!r} "
        f"sub_asset_class={sub_asset_class!r}"
    )
