"""Phase 24 — study renderers (markdown + xlsx). Synthetic fixtures only."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from aa_model.entity import (
    EntityFixture,
    EntityPolicyConfig,
    export_study_xlsx,
    load_entity_fixture,
    load_entity_policy,
    render_study_markdown,
)

_FIXTURE_REL = "data/fixtures/entities/entity_synth_a.yaml"
_POLICY_REL = "data/fixtures/entities/entity_synth_a_policy.yaml"


@pytest.fixture
def synth_fixture(repo_root: Path) -> EntityFixture:
    return load_entity_fixture(repo_root / _FIXTURE_REL)


@pytest.fixture
def synth_policy(repo_root: Path) -> EntityPolicyConfig:
    return load_entity_policy(repo_root / _POLICY_REL)


# ---- markdown --------------------------------------------------------------


def test_markdown_has_all_sections(synth_fixture, synth_policy) -> None:
    md = render_study_markdown(synth_fixture, policy=synth_policy)
    for heading in (
        "# Asset Allocation Study — entity_synth_a",
        "## Summary — portfolio at a glance",
        "## Balance-sheet segmentation",
        "## Allocation vs. strategic target",
        "## Holdings detail",
        "## Private equity & alternatives",
        "## Liquidity by redemption tier",
        "## Burn rate",
        "## Cash flow & runway",
        "## Liquidity projection",
        "## Custodian reconciliation",
        "## Notes & sources",
    ):
        assert heading in md, f"missing section: {heading!r}"
    # a couple of known formatted values
    assert "$100,000,000.00" in md  # total NAV
    assert "$40,000,000.00" in md  # investable base


def test_markdown_deterministic(synth_fixture, synth_policy) -> None:
    a = render_study_markdown(synth_fixture, policy=synth_policy)
    b = render_study_markdown(synth_fixture, policy=synth_policy)
    assert a == b


def test_markdown_without_policy_skips_allocation(synth_fixture) -> None:
    md = render_study_markdown(synth_fixture)
    assert "## Allocation vs. strategic target" not in md
    assert "## Summary — portfolio at a glance" in md  # other sections still render


# ---- xlsx ------------------------------------------------------------------


def test_xlsx_export(synth_fixture, synth_policy, tmp_path: Path) -> None:
    out = export_study_xlsx(synth_fixture, tmp_path / "study.xlsx", policy=synth_policy)
    assert out.exists()
    wb = openpyxl.load_workbook(out, data_only=True)
    for name in (
        "Summary",
        "Balance Sheet",
        "Allocation vs Target",
        "Holdings",
        "PE & Alternatives",
        "Liquidity",
        "Burn Rate",
        "Cash Flow",
        "Liquidity Projection",
        "Custodian Recon",
        "Notes",
    ):
        assert name in wb.sheetnames, f"missing sheet: {name}"
    # Summary: first data row is total NAV = 100,000,000
    ws = wb["Summary"]
    assert ws.cell(2, 1).value == "Total balance-sheet NAV"
    assert ws.cell(2, 2).value == 100000000.0
    wb.close()


def test_xlsx_without_policy_omits_allocation(synth_fixture, tmp_path: Path) -> None:
    out = export_study_xlsx(synth_fixture, tmp_path / "study2.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "Allocation vs Target" not in wb.sheetnames
    assert "Summary" in wb.sheetnames
    wb.close()


def test_empty_fixture_renders_minimal() -> None:
    fx = EntityFixture.model_validate(
        {"fixture_version": "t", "entity_id": "e1", "as_of_date": "2026-04-30"}
    )
    md = render_study_markdown(fx)
    assert "# Asset Allocation Study — e1" in md
    assert "## Notes & sources" in md
    # data-driven sections absent
    assert "## Burn rate" not in md
    assert "## Private equity" not in md
