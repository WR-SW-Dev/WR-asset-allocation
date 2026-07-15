"""Phase 26 — purpose (goals-based) allocation: config validation, holding→
purpose resolution, lens, renderer gating, CLI. Synthetic fixtures only —
per the design lock, the real workbook oracle runs local/gitignored.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from aa_model.entity import (
    EntityPurposePolicyConfig,
    load_entity_purpose_policy,
)
from pydantic import ValidationError

D = Decimal


def _band(purpose: str, target: str, lo: str = "0", hi: str = "0") -> dict:
    return {
        "purpose": purpose,
        "target": target,
        "lower_band_pp": lo,
        "upper_band_pp": hi,
    }


def _config(**overrides) -> dict:
    cfg = {
        "purpose_policy_version": "synth_purpose_v1",
        "entity_id": "entity_synth_a",
        "bands": {
            "liquidity": _band("liquidity", "0.05", "0.03", "0.10"),
            "stability": _band("stability", "0.10", "0.05", "0.05"),
            "income": _band("income", "0.34", "0.10", "0.10"),
            "growth": _band("growth", "0.45", "0.10", "0.10"),
            "aggressive_growth": _band("aggressive_growth", "0.05", "0.05", "0.05"),
            "community": _band("community", "0.01", "0.01", "0.03"),
        },
        "assignments": {},
        "default_by_policy_class": {
            "cash_and_cash_alts": "liquidity",
            "fixed_income": "income",
            "equity": "growth",
            "real_estate": "growth",
            "private_equity": "aggressive_growth",
            "absolute_return": "hedge",
            "re_opco_stabilized": "growth",
        },
    }
    cfg.update(overrides)
    return cfg


# ---- group 1: config validation --------------------------------------------


def test_config_valid_and_derived_bounds() -> None:
    cfg = EntityPurposePolicyConfig.model_validate(_config())
    liq = cfg.bands["liquidity"]
    assert liq.min_pct == D("0.02")
    assert liq.max_pct == D("0.15")
    # floor at zero: community 1% target with 1pp lower band
    com = cfg.bands["community"]
    assert com.min_pct == D("0")
    assert com.max_pct == D("0.04")


def test_config_targets_must_sum_to_one() -> None:
    bad = _config()
    bad["bands"]["growth"] = _band("growth", "0.30", "0.10", "0.10")  # sum 0.85
    with pytest.raises(ValidationError, match="sum to 1.0"):
        EntityPurposePolicyConfig.model_validate(bad)


def test_config_negative_band_rejected() -> None:
    bad = _config()
    bad["bands"]["income"] = _band("income", "0.34", "-0.01", "0.10")
    with pytest.raises(ValidationError, match="finite and >= 0"):
        EntityPurposePolicyConfig.model_validate(bad)


def test_config_band_key_purpose_mismatch_rejected() -> None:
    bad = _config()
    bad["bands"]["income"] = _band("hedge", "0.34")
    with pytest.raises(ValidationError, match="must match its dict key"):
        EntityPurposePolicyConfig.model_validate(bad)


def test_config_unknown_purpose_rejected() -> None:
    bad = _config()
    bad["bands"]["speculation"] = _band("growth", "0.00")
    with pytest.raises(ValidationError):
        EntityPurposePolicyConfig.model_validate(bad)


def test_config_version_must_be_url_safe() -> None:
    with pytest.raises(ValidationError, match="URL-safe"):
        EntityPurposePolicyConfig.model_validate(_config(purpose_policy_version="not ok/version"))


def test_config_strict_extra_keys_rejected() -> None:
    with pytest.raises(ValidationError):
        EntityPurposePolicyConfig.model_validate(_config(unexpected_key=1))


def test_config_empty_bands_rejected() -> None:
    with pytest.raises(ValidationError, match="non-empty"):
        EntityPurposePolicyConfig.model_validate(_config(bands={}))


def test_loader_yaml_roundtrip(tmp_path: Path) -> None:
    import yaml

    path = tmp_path / "purpose_policy.yaml"
    path.write_text(yaml.safe_dump(_config()), encoding="utf-8")
    cfg = load_entity_purpose_policy(path)
    assert cfg.purpose_policy_version == "synth_purpose_v1"
    assert cfg.bands["growth"].target == D("0.45")


def test_loader_rejects_unknown_extension(tmp_path: Path) -> None:
    path = tmp_path / "purpose_policy.txt"
    path.write_text("{}", encoding="utf-8")
    with pytest.raises(ValueError, match="unsupported purpose-policy extension"):
        load_entity_purpose_policy(path)


# ---- groups 2–4: resolution + lens ------------------------------------------

from aa_model.entity import (  # noqa: E402
    EntityFixture,
    purpose_allocation_lens,
    resolve_holding_purpose,
)


def _full_fixture() -> EntityFixture:
    """Synthetic fixture whose holdings fully cover the investable base of
    1000: cash 200 (150 money-market + 50 short-duration ETF — the class that
    splits across two purposes), fixed_income 340, equity 300, real_estate
    100, private_equity 50, absolute_return 10; structural 500 outside.
    """
    return EntityFixture.model_validate(
        {
            "fixture_version": "synth_purpose_v1",
            "entity_id": "entity_synth_a",
            "as_of_date": "2026-07-15",
            "accounts": [],
            "segments": [
                {
                    "segment_key": "inv_cash",
                    "segment": "investable",
                    "policy_class": "cash_and_cash_alts",
                    "amount_usd": "200.00",
                },
                {
                    "segment_key": "inv_fi",
                    "segment": "investable",
                    "policy_class": "fixed_income",
                    "amount_usd": "340.00",
                },
                {
                    "segment_key": "inv_eq",
                    "segment": "investable",
                    "policy_class": "equity",
                    "amount_usd": "300.00",
                },
                {
                    "segment_key": "inv_re",
                    "segment": "investable",
                    "policy_class": "real_estate",
                    "amount_usd": "100.00",
                },
                {
                    "segment_key": "inv_pe",
                    "segment": "investable",
                    "policy_class": "private_equity",
                    "amount_usd": "50.00",
                },
                {
                    "segment_key": "inv_ar",
                    "segment": "investable",
                    "policy_class": "absolute_return",
                    "amount_usd": "10.00",
                },
                {"segment_key": "str_home", "segment": "personal_use", "amount_usd": "500.00"},
            ],
            "holdings": [
                {
                    "holding_key": "mmkt_fund",
                    "account_id": "acct",
                    "policy_class": "cash_and_cash_alts",
                    "market_value_usd": "150.00",
                },
                {
                    "holding_key": "short_dur_etf",
                    "account_id": "acct",
                    "policy_class": "cash_and_cash_alts",
                    "market_value_usd": "50.00",
                },
                {
                    "holding_key": "bond_fund",
                    "account_id": "acct",
                    "policy_class": "fixed_income",
                    "market_value_usd": "340.00",
                },
                {
                    "holding_key": "stock_fund",
                    "account_id": "acct",
                    "policy_class": "equity",
                    "market_value_usd": "300.00",
                },
                {
                    "holding_key": "re_lp",
                    "account_id": "acct",
                    "policy_class": "real_estate",
                    "market_value_usd": "100.00",
                },
                {
                    "holding_key": "pe_fund",
                    "account_id": "acct",
                    "policy_class": "private_equity",
                    "market_value_usd": "50.00",
                },
                {
                    "holding_key": "hedge_fund",
                    "account_id": "acct",
                    "policy_class": "absolute_return",
                    "market_value_usd": "10.00",
                },
            ],
        }
    )


def _purpose_policy(**overrides) -> EntityPurposePolicyConfig:
    cfg = _config()
    # split the cash class: explicit assignment sends the short-duration ETF
    # to stability while the class default sends the rest to liquidity.
    cfg["assignments"] = {"short_dur_etf": "stability"}
    cfg["bands"] = {
        "liquidity": _band("liquidity", "0.05", "0.03", "0.10"),
        "stability": _band("stability", "0.10", "0.05", "0.05"),
        "income": _band("income", "0.34", "0.10", "0.10"),
        "growth": _band("growth", "0.40", "0.10", "0.10"),
        "aggressive_growth": _band("aggressive_growth", "0.05", "0.05", "0.05"),
        "hedge": _band("hedge", "0.05", "0.05", "0.00"),
        "community": _band("community", "0.01", "0.01", "0.03"),
    }
    cfg.update(overrides)
    return EntityPurposePolicyConfig.model_validate(cfg)


def test_resolution_assignment_overrides_class_default() -> None:
    fx = _full_fixture()
    pp = _purpose_policy()
    short_dur = next(h for h in fx.holdings if h.holding_key == "short_dur_etf")
    mmkt = next(h for h in fx.holdings if h.holding_key == "mmkt_fund")
    assert resolve_holding_purpose(short_dur, pp) == "stability"
    assert resolve_holding_purpose(mmkt, pp) == "liquidity"


def test_resolution_unmapped_holding_raises() -> None:
    fx = _full_fixture()
    pp = _purpose_policy(default_by_policy_class={"cash_and_cash_alts": "liquidity"})
    with pytest.raises(ValueError, match="resolves to no purpose"):
        purpose_allocation_lens(fx, pp)


def test_resolution_stale_assignment_raises() -> None:
    fx = _full_fixture()
    pp = _purpose_policy(assignments={"short_dur_etf": "stability", "gone_fund": "hedge"})
    with pytest.raises(ValueError, match="stale"):
        purpose_allocation_lens(fx, pp)


def test_lens_entity_mismatch_raises() -> None:
    fx = _full_fixture()
    pp = _purpose_policy(entity_id="entity_other")
    with pytest.raises(ValueError, match="does not match"):
        purpose_allocation_lens(fx, pp)


def test_lens_statuses_and_arithmetic() -> None:
    fx = _full_fixture()
    pp = _purpose_policy()
    lens = purpose_allocation_lens(fx, pp)
    assert lens.investable_base_usd == D("1000.00")
    rows = {r.purpose: r for r in lens.rows}
    assert [r.purpose for r in lens.rows] == [
        "liquidity",
        "stability",
        "income",
        "growth",
        "aggressive_growth",
        "hedge",
        "community",
    ]
    # liquidity: 150/1000 = 15% vs band [2%, 15%] — inclusive upper edge
    assert rows["liquidity"].current_pct == D("0.15")
    assert rows["liquidity"].status == "in_band"
    # stability: 50/1000 = 5% vs band [5%, 15%] — inclusive lower edge
    assert rows["stability"].current_pct == D("0.05")
    assert rows["stability"].status == "in_band"
    # income: 340/1000 = 34% on target, in band; to-target = 0
    assert rows["income"].status == "in_band"
    assert rows["income"].to_target_usd == D("0.00")
    # growth: (300+100)/1000 = 40% on a 40% target
    assert rows["growth"].current_usd == D("400.00")
    assert rows["growth"].variance_pp == D("0")
    # hedge: 10/1000 = 1% vs band [0%, 5%] → in band despite −4pp variance
    assert rows["hedge"].variance_pp == D("-4.00")
    assert rows["hedge"].status == "in_band"
    # community: empty, 1% target, lower band floors at 0 → in band
    assert rows["community"].current_usd == D("0")
    assert rows["community"].status == "in_band"
    assert rows["community"].to_target_usd == D("10.00")
    # signed to-target arithmetic: liquidity 15% vs 5% target → trim 100
    assert rows["liquidity"].to_target_usd == D("-100.00")


def test_lens_below_and_above_band() -> None:
    fx = _full_fixture()
    pp = _purpose_policy()
    # tighten: stability min 6% → 5% current is below band;
    # liquidity max 14% → 15% current is above band
    cfg = _config()
    cfg["assignments"] = {"short_dur_etf": "stability"}
    cfg["bands"] = {
        "liquidity": _band("liquidity", "0.05", "0.03", "0.09"),
        "stability": _band("stability", "0.10", "0.04", "0.05"),
        "income": _band("income", "0.34", "0.10", "0.10"),
        "growth": _band("growth", "0.40", "0.10", "0.10"),
        "aggressive_growth": _band("aggressive_growth", "0.05", "0.05", "0.05"),
        "hedge": _band("hedge", "0.05", "0.05", "0.00"),
        "community": _band("community", "0.01", "0.01", "0.03"),
    }
    pp = EntityPurposePolicyConfig.model_validate(cfg)
    rows = {r.purpose: r for r in purpose_allocation_lens(fx, pp).rows}
    assert rows["liquidity"].status == "above_band"  # 15% > 14%
    assert rows["stability"].status == "below_band"  # 5% < 6%
    # empty purpose with positive min reports below_band
    cfg["bands"]["community"] = _band("community", "0.01", "0.00", "0.03")
    pp2 = EntityPurposePolicyConfig.model_validate(cfg)
    rows2 = {r.purpose: r for r in purpose_allocation_lens(fx, pp2).rows}
    assert rows2["community"].status == "below_band"


def test_lens_purpose_buckets_cover_base() -> None:
    fx = _full_fixture()
    pp = _purpose_policy()
    lens = purpose_allocation_lens(fx, pp)
    assert sum((r.current_usd for r in lens.rows), D("0")) == lens.investable_base_usd


def test_lens_partial_holdings_coverage_raises() -> None:
    data = _full_fixture().model_dump(mode="json")
    data["holdings"] = data["holdings"][:3]  # cover only part of the base
    fx = EntityFixture.model_validate(data)
    with pytest.raises(ValueError, match="do not cover the investable base"):
        purpose_allocation_lens(fx, _purpose_policy())


# ---- groups 5–6: renderers + CLI --------------------------------------------

import yaml as _yaml  # noqa: E402
from aa_model.entity import export_study_xlsx, render_study_markdown  # noqa: E402
from aa_model.entity.cli import main as cli_main  # noqa: E402


def test_render_markdown_gating() -> None:
    fx = _full_fixture()
    without = render_study_markdown(fx)
    assert "Purpose allocation" not in without
    with_pp = render_study_markdown(fx, purpose_policy=_purpose_policy())
    assert "## Purpose allocation (goals-based)" in with_pp
    assert "| liquidity |" in with_pp
    assert "in_band" in with_pp
    # section order: purpose section follows the allocation-vs-target position
    assert with_pp.index("## Purpose allocation") < with_pp.index("## Holdings detail")


def test_render_markdown_without_purpose_is_unchanged_by_feature() -> None:
    # the no-purpose render must not differ in any way when the argument is
    # omitted vs explicitly None (the byte-identical legacy contract)
    fx = _full_fixture()
    assert render_study_markdown(fx) == render_study_markdown(fx, purpose_policy=None)


def test_export_xlsx_purpose_sheet(tmp_path: Path) -> None:
    import openpyxl

    fx = _full_fixture()
    out = tmp_path / "study.xlsx"
    export_study_xlsx(fx, out, purpose_policy=_purpose_policy())
    wb = openpyxl.load_workbook(out)
    assert "Purpose Allocation" in wb.sheetnames
    ws = wb["Purpose Allocation"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Purpose" and header[-1] == "Status"
    purposes = [row[0].value for row in ws.iter_rows(min_row=2)]
    assert purposes == [
        "liquidity",
        "stability",
        "income",
        "growth",
        "aggressive_growth",
        "hedge",
        "community",
    ]
    # and gated off without the policy
    out2 = tmp_path / "study2.xlsx"
    export_study_xlsx(fx, out2)
    assert "Purpose Allocation" not in openpyxl.load_workbook(out2).sheetnames


def test_render_entity_mismatch_raises() -> None:
    fx = _full_fixture()
    with pytest.raises(ValueError, match="does not match"):
        render_study_markdown(fx, purpose_policy=_purpose_policy(entity_id="entity_other"))


def _write_purpose_fixture_and_policy(tmp_path: Path) -> tuple[Path, Path]:
    fx_path = tmp_path / "fixture.yaml"
    fx_path.write_text(_yaml.safe_dump(_full_fixture().model_dump(mode="json")), encoding="utf-8")
    pp_cfg = _config()
    pp_cfg["assignments"] = {"short_dur_etf": "stability"}
    pp_cfg["bands"] = {
        k: _band(k, t, lo, hi)
        for k, (t, lo, hi) in {
            "liquidity": ("0.05", "0.03", "0.10"),
            "stability": ("0.10", "0.05", "0.05"),
            "income": ("0.34", "0.10", "0.10"),
            "growth": ("0.40", "0.10", "0.10"),
            "aggressive_growth": ("0.05", "0.05", "0.05"),
            "hedge": ("0.05", "0.05", "0.00"),
            "community": ("0.01", "0.01", "0.03"),
        }.items()
    }
    pp_path = tmp_path / "purpose_policy.yaml"
    pp_path.write_text(_yaml.safe_dump(pp_cfg), encoding="utf-8")
    return fx_path, pp_path


def test_cli_purpose_policy_happy_path(tmp_path: Path) -> None:
    import json

    fx_path, pp_path = _write_purpose_fixture_and_policy(tmp_path)
    out = tmp_path / "out"
    rc = cli_main(
        [
            "--fixture",
            str(fx_path),
            "--purpose-policy",
            str(pp_path),
            "--out",
            str(out),
        ]
    )
    assert rc == 0
    md = (out / "study.md").read_text(encoding="utf-8")
    assert "## Purpose allocation (goals-based)" in md
    manifest = json.loads((out / "manifest.json").read_text())
    assert manifest["purpose_policy_version"] == "synth_purpose_v1"
    assert manifest["policy_version"] is None


def test_cli_purpose_policy_missing_file(tmp_path: Path) -> None:
    fx_path, _ = _write_purpose_fixture_and_policy(tmp_path)
    with pytest.raises(FileNotFoundError):
        cli_main(
            [
                "--fixture",
                str(fx_path),
                "--purpose-policy",
                str(tmp_path / "nope.yaml"),
                "--out",
                str(tmp_path / "out2"),
            ]
        )
